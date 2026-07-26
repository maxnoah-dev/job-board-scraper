"""Unit tests for the Excel exporter (reporting/excel_exporter.py)."""

from __future__ import annotations

import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from job_board_scraper.reporting.excel_exporter import (
    COLUMNS,
    DEFAULT_SHEET_NAME,
    ExcelExporter,
    ExcelExportOptions,
    _easy_apply_from_raw,
    _parse_salary_bucket,
)

# ---------------------------------------------------------------------------
# Salary bucket parser
# ---------------------------------------------------------------------------


class TestParseSalaryBucket:
    @pytest.mark.parametrize(
        "salary, expected",
        [
            ("$250K/yr", "250+"),
            ("$275,000/yr", "250+"),
            ("$240K/yr", "200+"),
            ("$199,999/yr", "150+"),
            ("$120K/yr", ""),
            ("", ""),
            (None, ""),
            ("$300K-$400K/yr", "250+"),
        ],
    )
    def test_parse_salary_bucket(self, salary: str | None, expected: str) -> None:
        assert _parse_salary_bucket(salary) == expected


# ---------------------------------------------------------------------------
# Easy Apply parser
# ---------------------------------------------------------------------------


class TestEasyApply:
    def test_true_bool_returns_yes(self) -> None:
        assert _easy_apply_from_raw({"easy_apply": True}) == "Yes"

    def test_false_bool_returns_blank(self) -> None:
        assert _easy_apply_from_raw({"easy_apply": False}) == ""

    def test_string_yes(self) -> None:
        assert _easy_apply_from_raw({"easy_apply": "yes"}) == "Yes"

    def test_missing(self) -> None:
        assert _easy_apply_from_raw({}) == ""
        assert _easy_apply_from_raw(None) == ""


# ---------------------------------------------------------------------------
# Exporter integration
# ---------------------------------------------------------------------------


def _sample_job() -> dict:
    return {
        "company_name": "OPSWAT",
        "company_slug": "opswat",
        "title": "Senior Backend Engineer",
        "title_vi": "Kỹ sư Backend cao cấp",
        "url": "https://jobs.opswat.com/positions/12345",
        "location": "Ho Chi Minh City, Vietnam",
        "status": "open",
        "salary_raw": "$180K/yr - $240K/yr",
        "raw_data": {"easy_apply": True},
    }


class TestExcelExporter:
    @pytest.mark.asyncio
    async def test_export_creates_new_workbook_with_17_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "out.xlsx"
            exporter = ExcelExporter(output_path=out_path)
            options = ExcelExportOptions(
                use_template_copy=False, sheet_name="Test"
            )
            result = await exporter.export([_sample_job()], options=options)
            assert result == out_path
            assert out_path.exists()

            wb = openpyxl.load_workbook(out_path)
            assert "Test" in wb.sheetnames
            ws = wb["Test"]
            assert ws.cell(row=1, column=1).value == COLUMNS[0]
            assert ws.cell(row=1, column=len(COLUMNS)).value == COLUMNS[-1]
            assert len(COLUMNS) == 17
            assert ws.cell(row=2, column=1).value == "OPSWAT"
            assert ws.cell(row=2, column=2).value == "Senior Backend Engineer"
            assert ws.cell(row=2, column=12).value == "Kỹ sư Backend cao cấp"
            assert ws.cell(row=2, column=16).value == "Yes"
            assert ws.cell(row=2, column=17).value == "150+"

    @pytest.mark.asyncio
    async def test_export_uses_template_when_provided(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Existing"
            ws["A1"] = "Original"
            wb.save(template_path)

            out_path = Path(tmpdir) / "out.xlsx"
            exporter = ExcelExporter(
                output_path=out_path, template_path=template_path
            )
            options = ExcelExportOptions(
                use_template_copy=True, sheet_name=DEFAULT_SHEET_NAME
            )
            await exporter.export([_sample_job()], options=options)
            wb2 = openpyxl.load_workbook(out_path)
            assert "Existing" in wb2.sheetnames
            assert DEFAULT_SHEET_NAME in wb2.sheetnames

    @pytest.mark.asyncio
    async def test_export_is_idempotent_when_rerun(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "out.xlsx"
            exporter = ExcelExporter(output_path=out_path)
            opts = ExcelExportOptions(
                use_template_copy=False, sheet_name=DEFAULT_SHEET_NAME
            )
            await exporter.export([_sample_job()], options=opts)
            await exporter.export([_sample_job(), _sample_job()], options=opts)
            wb = openpyxl.load_workbook(out_path)
            assert wb[DEFAULT_SHEET_NAME].max_row == 3  # header + 2 rows

    @pytest.mark.asyncio
    async def test_date_column_uses_today(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "out.xlsx"
            exporter = ExcelExporter(output_path=out_path)
            opts = ExcelExportOptions(use_template_copy=False)
            await exporter.export([_sample_job()], options=opts)
            wb = openpyxl.load_workbook(out_path)
            ws = wb[DEFAULT_SHEET_NAME]
            date_cell = ws.cell(row=2, column=8).value
            assert isinstance(date_cell, (datetime, type(ws.cell(row=2, column=1).value)))
