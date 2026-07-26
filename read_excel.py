"""Inspect the Excel template."""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402

TEMPLATE = Path(r"d:\Sources\job-board-scraper\docs\Apply Job in US.xlsx")

wb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== Sheet: {name} ===")
    print("max_row:", ws.max_row, "max_col:", ws.max_column)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        print(f"row {i}: {row}")
