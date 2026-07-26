"""Excel exporter for the job board scraper.

Mirrors the column layout of the ``Job in California`` sheet in
``docs/Apply Job in US.xlsx``:

    1. Company name
    2. Title job
    3. Link job
    4. Salary
    5. Location
    6. Loại job         (workflow)
    7. Status            (defaults to "Chưa apply")
    8. Date              (today by default)
    9. Người Apply       (workflow, blank)
   10. Kiểm tra          (workflow, blank)
   11. Ghi chú           (workflow, blank)
   12. Title job vietsub (from Job.title_vi or Vilao on-demand)
   13. TN-visa           (workflow, blank)
   14. Job Refer         (workflow, blank)
   15. Foundation        (workflow, blank)
   16. Easy Apply        ("Yes" if raw_data.easy_apply is true)
   17. Salary bucket     ("250+", "200+", "150+" — parsed from Salary)

The exporter writes a new sheet to a copy of the template workbook so
existing sheets are preserved.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    pass


DEFAULT_SHEET_NAME = "From Scraper"

COLUMNS: list[str] = [
    "Company name",
    "Title job",
    "Link job",
    "Salary",
    "Location",
    "Loại job",
    "Status",
    "Date",
    "Người Apply",
    "Kiểm tra",
    "Ghi chú",
    "Title job vietsub",
    "TN-visa",
    "Job Refer",
    "Foundation",
    "Easy Apply",
    "Salary bucket",
]

_SALARY_NUMBER_RE = re.compile(r"(\d{2,3})")


def _parse_salary_bucket(salary: str | None) -> str:
    """Return ``"250+"``, ``"200+"``, ``"150+"`` or empty based on $K.

    The lookup is a simple upper bound: any value >= 250 -> "250+",
    >= 200 -> "200+", >= 150 -> "150+".
    """
    if not salary:
        return ""
    match = _SALARY_NUMBER_RE.search(salary.replace(",", ""))
    if not match:
        return ""
    try:
        value = int(match.group(1))
    except ValueError:
        return ""
    if value >= 250:
        return "250+"
    if value >= 200:
        return "200+"
    if value >= 150:
        return "150+"
    return ""


def _easy_apply_from_raw(raw_data: dict[str, Any] | None) -> str:
    """Return ``"Yes"`` when the raw job advertises Easy Apply."""
    if not raw_data:
        return ""
    easy = raw_data.get("easy_apply")
    if isinstance(easy, bool):
        return "Yes" if easy else ""
    if isinstance(easy, str) and easy.lower() in {"yes", "true", "1"}:
        return "Yes"
    return ""


@dataclass
class ExcelExportOptions:
    """Tuning knobs for the Excel exporter."""

    sheet_name: str = DEFAULT_SHEET_NAME
    include_closed: bool = False
    default_status: str = "Chưa apply"
    use_template_copy: bool = True
    template_path: Path | None = None
    output_path: Path | None = None


class ExcelExporter:
    """Render jobs to the Vietnamese-first Excel template."""

    def __init__(
        self,
        output_path: Path | str,
        template_path: Path | str | None = None,
    ) -> None:
        self._output_path = Path(output_path)
        self._template_path = (
            Path(template_path) if template_path else None
        )

    async def export(
        self,
        jobs: Iterable[dict[str, Any]],
        options: ExcelExportOptions | None = None,
    ) -> Path:
        """Write the workbook atomically and return the final path."""
        options = options or ExcelExportOptions()
        rows = [self._job_to_row(job, options) for job in jobs]
        return self._write(rows, options)

    # ─── Internals ──────────────────────────────────────────────────────────
    @staticmethod
    def _job_to_row(
        job: dict[str, Any], options: ExcelExportOptions
    ) -> list[Any]:
        raw_data = job.get("raw_data") or {}
        salary = job.get("salary_raw") or job.get("salary") or raw_data.get("salary") or ""
        title_vi = job.get("title_vi") or ""
        return [
            job.get("company_name") or job.get("company_slug") or "",
            job.get("title") or "",
            job.get("url") or job.get("job_url") or "",
            salary,
            job.get("location") or "Remote",
            "",  # Loại job — workflow
            options.default_status,
            datetime.now(UTC).date(),
            "",  # Người Apply
            "",  # Kiểm tra
            "",  # Ghi chú
            title_vi,
            "",  # TN-visa
            "",  # Job Refer
            "",  # Foundation
            _easy_apply_from_raw(raw_data if isinstance(raw_data, dict) else None),
            _parse_salary_bucket(salary if isinstance(salary, str) else None),
        ]

    def _write(self, rows: list[list[Any]], options: ExcelExportOptions) -> Path:
        output_path = options.output_path or self._output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        target = self._prepare_workbook(output_path, options)
        try:
            self._populate_sheet(target, options.sheet_name, rows)
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".xlsx", dir=str(output_path.parent)
            ) as tmp:
                tmp_path = Path(tmp.name)
            target.save(tmp_path)
            os.replace(tmp_path, output_path)
        except Exception:
            raise
        return output_path

    def _maybe_load_template(self) -> Workbook | None:
        """Load the Excel template if it exists; otherwise return ``None``."""
        template = self._template_path
        if template is None:
            return None
        if not template.exists():
            return None
        return openpyxl.load_workbook(template)

    def _prepare_workbook(
        self, output_path: Path, options: ExcelExportOptions
    ) -> Workbook:
        """Return a workbook to populate.

        When ``use_template_copy`` is True and the template exists, copy the
        template to a sibling temp file so the original is left untouched.
        Otherwise build a fresh workbook.
        """
        if not options.use_template_copy:
            wb = openpyxl.Workbook()
            default_ws = wb.active
            wb.remove(default_ws)
            return wb
        template = self._template_path
        if template is not None and template.exists():
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".xlsx", dir=str(output_path.parent)
            ) as tmp:
                tmp_path = Path(tmp.name)
            shutil.copyfile(template, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            # Drop any prior "From Scraper" sheet so reruns are idempotent.
            if options.sheet_name in wb.sheetnames:
                del wb[options.sheet_name]
            return wb
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        return wb

    @staticmethod
    def _populate_sheet(
        workbook: Workbook, sheet_name: str, rows: list[list[Any]]
    ) -> None:
        ws: Worksheet = workbook.create_sheet(title=sheet_name)  # type: ignore[assignment]
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
        )
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        # Auto-width approximation.
        for col_idx, header in enumerate(COLUMNS, start=1):
            max_len = len(header)
            for row in rows:
                cell_value = row[col_idx - 1]
                if cell_value is None:
                    continue
                max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(15, max_len + 2), 60
            )
        ws.freeze_panes = "A2"


__all__ = [
    "ExcelExporter",
    "ExcelExportOptions",
    "COLUMNS",
    "DEFAULT_SHEET_NAME",
]
